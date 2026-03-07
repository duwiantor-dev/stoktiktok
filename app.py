import re
from io import BytesIO
from typing import Dict, Tuple, List, Optional, Set

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# CONFIG / LIMITS
# ============================================================
MAX_MASS_FILES = 50
MAX_TOTAL_UPLOAD_MB = 200
AREA_CODE_RE = re.compile(r"^\d+[A-Z]$")


# ============================================================
# HELPERS
# ============================================================
def _norm_str(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def norm_sku(v) -> str:
    s = _norm_str(v).upper()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"\s+", "", s)
    return s


def split_sku_addons(sku_full: str) -> Tuple[str, List[str]]:
    parts = [p for p in _norm_str(sku_full).split("+") if p.strip()]
    if not parts:
        return "", []
    return parts[0].strip(), [p.strip() for p in parts[1:]]


def to_int_or_none(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if v != v:
            return None
        return int(round(v))
    s = _norm_str(v)
    if not s:
        return None
    digits = re.findall(r"\d+", s)
    if not digits:
        return None
    return int("".join(digits))


def get_first_sheet(wb: openpyxl.Workbook) -> Worksheet:
    return wb[wb.sheetnames[0]]


def find_row_contains(ws: Worksheet, needle: str, scan_rows: int = 300) -> Optional[int]:
    needle_u = needle.upper()
    max_r = min(ws.max_row, scan_rows)
    for r in range(1, max_r + 1):
        for c in range(1, ws.max_column + 1):
            v = _norm_str(ws.cell(r, c).value).upper()
            if v and (needle_u == v or needle_u in v):
                return r
    return None


def get_merged_value(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row, col)
    if cell.value not in (None, ""):
        return cell.value
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col).value
    return None


def sheet_range_between(sheetnames: List[str], start: str, end: str) -> List[str]:
    up = [s.upper() for s in sheetnames]
    if start.upper() not in up or end.upper() not in up:
        raise ValueError(f"Sheet range tidak valid. Pastikan ada '{start}' dan '{end}'.")
    i0 = up.index(start.upper())
    i1 = up.index(end.upper())
    if i0 > i1:
        i0, i1 = i1, i0
    return sheetnames[i0:i1 + 1]


def norm_area_default(area_raw: str) -> str:
    a = _norm_str(area_raw).upper()
    if not a:
        return ""
    if a.endswith("A"):
        return a[:-1]
    return a


def parse_pairs_space(text: str) -> Set[Tuple[str, str]]:
    out = set()
    if not text:
        return out
    chunks = re.split(r"[;\n]+", text)
    for ch in chunks:
        ch = ch.strip()
        if not ch:
            continue
        parts = ch.split()
        if len(parts) < 2:
            continue
        toko = parts[0].strip().upper()
        area = norm_area_default(parts[1].strip().upper())
        if toko and area:
            out.add((toko, area))
    return out


def total_upload_size_mb(files: List) -> float:
    total = 0
    for f in files:
        try:
            total += len(f.getvalue())
        except Exception:
            pass
    return total / (1024 * 1024)


# ============================================================
# PRICE LIST PARSING
# ============================================================
def delete_coming_block_in_laptop(ws: Worksheet):
    r_start = find_row_contains(ws, "COMING", scan_rows=600)
    r_end = find_row_contains(ws, "END COMING", scan_rows=1200)
    if r_start and r_end and r_end >= r_start:
        ws.delete_rows(r_start, r_end - r_start + 1)


def find_header_row_by_exact(ws: Worksheet, header_text: str, scan_rows: int = 120) -> Optional[int]:
    target = header_text.strip().upper()
    max_r = min(ws.max_row, scan_rows)
    for r in range(1, max_r + 1):
        for c in range(1, ws.max_column + 1):
            v = _norm_str(ws.cell(r, c).value).strip().upper()
            if v == target:
                return r
    return None


def find_tot_col(ws: Worksheet, header_row_hint: int) -> Tuple[int, int]:
    for c in range(1, ws.max_column + 1):
        if _norm_str(ws.cell(header_row_hint, c).value).strip().upper() == "TOT":
            return header_row_hint, c

    for r in range(1, min(12, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_str(ws.cell(r, c).value).strip().upper() == "TOT":
                return r, c

    raise ValueError("Kolom 'TOT' tidak ketemu.")


def detect_area_and_toko_rows(ws: Worksheet, header_row: int) -> Tuple[int, int]:
    pairs = []
    if ws.max_row >= 4:
        pairs.append((3, 4))
    r1, r2 = header_row + 1, header_row + 2
    if ws.max_row >= r2:
        pairs.append((r1, r2))

    def score_area_row(r: int) -> int:
        cnt = 0
        for c in range(1, ws.max_column + 1):
            v = _norm_str(get_merged_value(ws, r, c)).upper()
            if v and AREA_CODE_RE.match(v):
                cnt += 1
        return cnt

    best = None
    best_score = -1
    for a, b in pairs:
        sa, sb = score_area_row(a), score_area_row(b)
        score = max(sa, sb)
        if score > best_score:
            best_score = score
            best = (a, b) if sa >= sb else (b, a)

    return best if best else (4, 3)


def build_stock_lookup_from_sheet(ws: Worksheet, sheet_name: str):
    header_row = find_header_row_by_exact(ws, "KODEBARANG", scan_rows=150)
    if header_row is None:
        header_row = find_header_row_by_exact(ws, "KODE BARANG", scan_rows=150)
    if header_row is None:
        raise ValueError(f"[{sheet_name}] Header 'KODEBARANG' tidak ketemu.")

    sku_col = None
    for c in range(1, ws.max_column + 1):
        v = _norm_str(ws.cell(header_row, c).value).strip().upper()
        if v in ("KODEBARANG", "KODE BARANG"):
            sku_col = c
            break

    if sku_col is None:
        raise ValueError(f"[{sheet_name}] Kolom 'KODEBARANG'/'KODE BARANG' tidak ditemukan.")

    header_row_used, tot_col = find_tot_col(ws, header_row)
    area_row, toko_row = detect_area_and_toko_rows(ws, header_row_used)

    sku_map: Dict[str, Dict] = {}
    tokos: Set[str] = set()
    areas: Set[str] = set()

    for r in range(header_row_used + 1, ws.max_row + 1):
        sku_raw = ws.cell(r, sku_col).value
        sku = _norm_str(sku_raw)
        if not sku:
            continue

        sku_key = norm_sku(sku)
        if sku_key in ("TOTAL", "KODEBARANG", "KODEBARANG."):
            continue

        tot_val = to_int_or_none(ws.cell(r, tot_col).value)
        by_toko_area = {}

        for c in range(tot_col + 1, ws.max_column + 1):
            area_raw = get_merged_value(ws, area_row, c)
            toko_raw = get_merged_value(ws, toko_row, c)

            area_s = _norm_str(area_raw).upper()
            toko_s = _norm_str(toko_raw).upper()

            if not area_s and not toko_s:
                continue

            if AREA_CODE_RE.match(toko_s) and not AREA_CODE_RE.match(area_s):
                area_s, toko_s = toko_s, area_s

            if not area_s or not toko_s:
                continue
            if not AREA_CODE_RE.match(area_s):
                continue

            area_n = norm_area_default(area_s)
            if not area_n:
                continue

            v = to_int_or_none(ws.cell(r, c).value)
            if v is None:
                continue

            tokos.add(toko_s)
            areas.add(area_n)
            by_toko_area[(toko_s, area_n)] = v

        sku_map[sku_key] = {"TOT": tot_val, "by_toko_area": by_toko_area}

    return sku_map, sorted(tokos), sorted(areas)


@st.cache_data(show_spinner=False)
def build_stock_lookup_from_pricelist_cached(pl_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(pl_bytes), data_only=True, read_only=False)

    for s in wb.sheetnames:
        if s.upper() == "LAPTOP":
            delete_coming_block_in_laptop(wb[s])
            break

    target_sheets = sheet_range_between(wb.sheetnames, "LAPTOP", "SER OTH CON")

    merged: Dict[str, Dict] = {}
    tokos_all: Set[str] = set()
    areas_all: Set[str] = set()

    for s in target_sheets:
        ws = wb[s]
        sku_map, tokos, areas = build_stock_lookup_from_sheet(ws, s)
        merged.update(sku_map)
        tokos_all |= set(tokos)
        areas_all |= set(areas)

    if not merged:
        raise ValueError("Pricelist terbaca, tapi lookup stok kosong.")

    return merged, sorted(tokos_all), sorted(areas_all)


# ============================================================
# TIKTOK LAYOUT
# ============================================================
def find_tiktok_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    HEADER_ROW = 3
    DATA_START_ROW = 6
    sku_col = None
    qty_col = None

    for c in range(1, ws.max_column + 1):
        v = _norm_str(ws.cell(HEADER_ROW, c).value).strip().upper()
        if v in ("SKU PENJUAL", "SELLER SKU"):
            sku_col = c
        if v == "KUANTITAS":
            qty_col = c

    if not sku_col:
        raise ValueError("Kolom SKU tidak ketemu. Pastikan header 'SKU Penjual' / 'Seller SKU' ada di row 3.")
    if not qty_col:
        raise ValueError("Kolom Kuantitas tidak ketemu. Pastikan header 'Kuantitas' ada di row 3.")

    return DATA_START_ROW, sku_col, qty_col


def find_tiktok_columns_readonly(ws) -> Tuple[int, int, int]:
    HEADER_ROW = 3
    DATA_START_ROW = 6
    sku_col = None
    qty_col = None

    row_vals = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
    for idx, val in enumerate(row_vals, start=1):
        v = _norm_str(val).strip().upper()
        if v in ("SKU PENJUAL", "SELLER SKU"):
            sku_col = idx
        if v == "KUANTITAS":
            qty_col = idx

    if not sku_col:
        raise ValueError("Kolom SKU tidak ketemu. Pastikan header 'SKU Penjual' / 'Seller SKU' ada di row 3.")
    if not qty_col:
        raise ValueError("Kolom Kuantitas tidak ketemu. Pastikan header 'Kuantitas' ada di row 3.")

    return DATA_START_ROW, sku_col, qty_col


def pick_stock_value(
    sku_full: str,
    stock_lookup: Dict[str, Dict],
    mode: str,
    chosen_tokos: Set[str],
    chosen_pairs: Set[Tuple[str, str]],
) -> Optional[int]:
    base, _ = split_sku_addons(sku_full)
    base_key = norm_sku(base)
    if not base_key or base_key not in stock_lookup:
        return None

    rec = stock_lookup[base_key]
    tot = rec.get("TOT")
    by_toko_area: Dict[Tuple[str, str], int] = rec.get("by_toko_area", {}) or {}

    if mode == "Stok Nasional (TOT)":
        return tot if tot is not None else None

    if mode == "Stok Area":
        if not chosen_tokos:
            return None
        s, hit = 0, False
        for (t, a), v in by_toko_area.items():
            if t in chosen_tokos:
                s += int(v)
                hit = True
        return s if hit else None

    if mode == "Stok Toko":
        if not chosen_pairs:
            return None
        s, hit = 0, False
        for key, v in by_toko_area.items():
            if key in chosen_pairs:
                s += int(v)
                hit = True
        return s if hit else None

    return None


# ============================================================
# FASTER PROCESSOR
# ============================================================
def collect_changed_rows_from_mass_file(
    file_name: str,
    file_bytes: bytes,
    stock_lookup: Dict[str, Dict],
    mode: str,
    chosen_tokos: Set[str],
    chosen_pairs: Set[Tuple[str, str]],
):
    """
    Baca file mass update dalam read_only mode, lalu ambil hanya row yang berubah.
    Return:
      changed_rows: List[List[object]]
      stats
    """
    stats = {
        "rows_scanned": 0,
        "rows_written": 0,
        "rows_unchanged": 0,
        "rows_unmatched": 0,
    }

    changed_rows = []

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]

    data_start, sku_col, qty_col = find_tiktok_columns_readonly(ws)

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        row_list = list(row)
        if not row_list:
            continue

        sku_full = _norm_str(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue

        stats["rows_scanned"] += 1

        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(
            sku_full=sku_full,
            stock_lookup=stock_lookup,
            mode=mode,
            chosen_tokos=chosen_tokos,
            chosen_pairs=chosen_pairs,
        )

        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue

        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue

        # pastikan panjang row cukup
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))

        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1

    wb.close()
    return changed_rows, stats


def write_output_from_template(
    template_bytes: bytes,
    changed_rows_all: List[List[object]],
) -> bytes:
    """
    Pakai file mass pertama sebagai template.
    Header/format atas dipertahankan, area data lama dibersihkan lalu diisi ulang.
    Tidak pakai insert_rows.
    """
    out_wb = openpyxl.load_workbook(BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)

    data_start, sku_col, qty_col = find_tiktok_columns_normal(out_ws)

    # hapus data lama
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)

    # tulis hasil baru langsung
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val

    buf = BytesIO()
    out_wb.save(buf)
    return buf.getvalue()


def process_mass_update_stock_tiktok_fast(
    mass_files: List[Tuple[str, bytes]],
    stock_lookup: Dict[str, Dict],
    mode: str,
    chosen_tokos: Set[str],
    chosen_pairs: Set[Tuple[str, str]],
    progress_slot=None,
) -> Tuple[bytes, pd.DataFrame, Dict[str, int]]:
    issues = []
    changed_rows_all: List[List[object]] = []
    stats_total = {
        "files_total": len(mass_files),
        "rows_scanned": 0,
        "rows_written": 0,
        "rows_unchanged": 0,
        "rows_unmatched": 0,
    }

    for idx, (fname, fbytes) in enumerate(mass_files, start=1):
        try:
            changed_rows, stats = collect_changed_rows_from_mass_file(
                file_name=fname,
                file_bytes=fbytes,
                stock_lookup=stock_lookup,
                mode=mode,
                chosen_tokos=chosen_tokos,
                chosen_pairs=chosen_pairs,
            )
            changed_rows_all.extend(changed_rows)

            for k in ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"):
                stats_total[k] += stats[k]

        except Exception as e:
            issues.append({"file": fname, "reason": f"Gagal proses file: {e}"})

        if progress_slot is not None:
            progress_slot.progress(idx / len(mass_files), text=f"Memproses file {idx}/{len(mass_files)}")

    if stats_total["rows_written"] == 0:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    out_bytes = write_output_from_template(mass_files[0][1], changed_rows_all)
    issues_df = pd.DataFrame(issues, columns=["file", "reason"])
    return out_bytes, issues_df, stats_total


# ============================================================
# UI
# ============================================================
st.set_page_config(page_title="Update Stok TikTok (Mass Update)", layout="wide")
st.title("Update Stok TikTok (Mass Update)")

# Session state init
if "stock_lookup" not in st.session_state:
    st.session_state.stock_lookup = None
if "tokos" not in st.session_state:
    st.session_state.tokos = []
if "areas" not in st.session_state:
    st.session_state.areas = []
if "result_bytes" not in st.session_state:
    st.session_state.result_bytes = None
if "issues_df" not in st.session_state:
    st.session_state.issues_df = None
if "stats" not in st.session_state:
    st.session_state.stats = None
if "last_result_name" not in st.session_state:
    st.session_state.last_result_name = "hasil_update_stok_tiktok.xlsx"

col1, col2 = st.columns(2)

with col1:
    mass_uploads = st.file_uploader(
        "Upload Mass Update TikTok (bisa banyak)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="mass_uploads",
    )

with col2:
    pl_upload = st.file_uploader(
        "Upload Pricelist (multi-sheet)",
        type=["xlsx", "XLSX"],
        accept_multiple_files=False,
        key="pl_upload",
    )

st.caption("Catatan: SKU yang mengandung '+ADDON' akan pakai stok BASE SKU (sebelum '+').")

# Guard upload size
if mass_uploads:
    if len(mass_uploads) > MAX_MASS_FILES:
        st.error(f"Maksimal {MAX_MASS_FILES} file mass update per proses.")
        st.stop()

    total_mb = total_upload_size_mb(mass_uploads)
    if total_mb > MAX_TOTAL_UPLOAD_MB:
        st.error(f"Total ukuran file terlalu besar: {total_mb:.1f} MB. Maksimal {MAX_TOTAL_UPLOAD_MB} MB.")
        st.stop()

load_btn = st.button("Load Data (Ambil daftar TOKO & AREA)", type="secondary", key="btn_load_data")

if load_btn:
    if not pl_upload:
        st.error("Upload Pricelist dulu.")
        st.stop()

    try:
        with st.spinner("Membaca pricelist..."):
            pl_bytes = pl_upload.getvalue()
            lookup, tokos, areas = build_stock_lookup_from_pricelist_cached(pl_bytes)
            st.session_state.stock_lookup = lookup
            st.session_state.tokos = tokos
            st.session_state.areas = areas

        st.success(f"OK. Ditemukan {len(tokos)} TOKO dan {len(areas)} AREA.")
    except Exception as e:
        st.error(f"Pricelist tidak valid: {e}")
        st.stop()

mode = st.radio(
    "Pilih sumber stok untuk update",
    options=["Stok Nasional (TOT)", "Stok Area", "Stok Toko"],
    horizontal=True,
    key="mode_stock_source",
)

chosen_tokos: Set[str] = set()
chosen_pairs: Set[Tuple[str, str]] = set()

if st.session_state.stock_lookup is None:
    st.info("Klik 'Load Data' dulu supaya daftar TOKO & AREA muncul.")
else:
    if mode == "Stok Area":
        chosen_tokos = set(
            st.multiselect(
                "Pilih TOKO (boleh banyak)",
                options=st.session_state.tokos,
                key="ms_tokos",
            )
        )
    elif mode == "Stok Toko":
        st.caption("Masukkan Kode Toko dan AREA. Contoh: RAM 2A ; JKT 3B")
        pairs_text = st.text_area(
            "PILIH BANYAK (pisah dengan enter atau ';')",
            value="",
            height=120,
            key="ta_pairs",
        )
        chosen_pairs = parse_pairs_space(pairs_text)

with st.expander("DEBUG (cek kolom & match SKU)", expanded=False):
    if mass_uploads:
        try:
            f0 = mass_uploads[0]
            wb = openpyxl.load_workbook(BytesIO(f0.getvalue()), read_only=True, data_only=False)
            ws = wb[wb.sheetnames[0]]
            ds, sc, qc = find_tiktok_columns_readonly(ws)

            st.write("Mass Update -> data_start:", ds, "| sku_col:", sc, "| qty_col:", qc)

            sample_mass = []
            for idx, row in enumerate(ws.iter_rows(min_row=ds, values_only=True), start=1):
                if idx > 30:
                    break
                sv = norm_sku(row[sc - 1] if len(row) >= sc else None)
                if sv:
                    sample_mass.append(sv)

            st.write("Sample SKU Mass Update:", sample_mass[:10])

            if st.session_state.stock_lookup is not None:
                pl_keys = list(st.session_state.stock_lookup.keys())
                st.write("Sample SKU Pricelist:", pl_keys[:10])
                inter = set(sample_mass) & set(pl_keys)
                st.write("Match count (sample 30 baris):", len(inter))

            wb.close()

        except Exception as e:
            st.write("DEBUG error:", str(e))
    else:
        st.write("Upload Mass Update dulu untuk debug.")

run = st.button("Proses Update Stok", type="primary", key="btn_run")

if run:
    if not mass_uploads:
        st.error("Mass Update wajib diupload.")
        st.stop()

    if st.session_state.stock_lookup is None:
        st.error("Klik 'Load Data' dulu.")
        st.stop()

    if mode == "Stok Area" and not chosen_tokos:
        st.error("Mode Stok Area: pilih minimal 1 TOKO.")
        st.stop()

    if mode == "Stok Toko" and not chosen_pairs:
        st.error("Mode Stok Toko: isi minimal 1 pasangan 'TOKO AREA'.")
        st.stop()

    try:
        progress = st.progress(0, text="Menyiapkan proses...")

        with st.spinner("Sedang proses update stok TikTok..."):
            mass_files = [(f.name, f.getvalue()) for f in mass_uploads]

            out_bytes, issues_df, stats = process_mass_update_stock_tiktok_fast(
                mass_files=mass_files,
                stock_lookup=st.session_state.stock_lookup,
                mode=mode,
                chosen_tokos=chosen_tokos,
                chosen_pairs=chosen_pairs,
                progress_slot=progress,
            )

            st.session_state.result_bytes = out_bytes
            st.session_state.issues_df = issues_df
            st.session_state.stats = stats
            st.session_state.last_result_name = "hasil_update_stok_tiktok.xlsx"

        progress.progress(1.0, text="Selesai")
        st.success("Selesai proses update stok.")

    except Exception as e:
        st.error(str(e))

if st.session_state.result_bytes is not None:
    st.subheader("Hasil Proses")

    if st.session_state.stats:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("File diproses", st.session_state.stats["files_total"])
        c2.metric("Baris discan", st.session_state.stats["rows_scanned"])
        c3.metric("Baris diupdate", st.session_state.stats["rows_written"])
        c4.metric("SKU tidak match", st.session_state.stats["rows_unmatched"])

    st.download_button(
        "Download hasil (XLSX)",
        data=st.session_state.result_bytes,
        file_name=st.session_state.last_result_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_result",
    )

    if st.session_state.issues_df is not None and len(st.session_state.issues_df) > 0:
        st.subheader("Issues Report")
        st.dataframe(st.session_state.issues_df, use_container_width=True)

if st.button("Reset hasil", key="btn_reset_result"):
    st.session_state.result_bytes = None
    st.session_state.issues_df = None
    st.session_state.stats = None
    st.success("Hasil proses di-reset.")
